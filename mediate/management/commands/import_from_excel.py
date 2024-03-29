from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

import os

from catalogues.models import Collection, Catalogue, Lot, Category
from items.models import Item, Edition, BookFormat
from persons.models import Place


class Command(BaseCommand):
    help = 'Imports text transcriptions'

    def add_arguments(self, parser):
        # Positional
        parser.add_argument('transcription', type=str, nargs='+', help='Text transcription file')

        # Optional
        parser.add_argument('-d', '--dry_run', action='store_true', help='Do not save to database.')
        parser.add_argument('-vv', '--verbose', action='store_true', help='Print the data.')

    @transaction.atomic
    def handle(self, *args, **kwargs):
        # Get the command line arguments
        transcription_files = kwargs['transcription']
        dry_run = kwargs["dry_run"]
        verbose = kwargs["verbose"]

        for transcription_file in transcription_files:
            # Get data from Excel into a list of dicts
            # where each dict represents a lot
            workbook = load_workbook(filename=transcription_file)
            worksheet = workbook[workbook.get_sheet_names()[0]]
            header = [cell.value for cell in worksheet['B1':'Y1'][0]]
            lots_data = []
            for row in worksheet.iter_rows(min_row=2, max_col=25, min_col=2):
                row_as_list = [cell.value for cell in row]
                lots_data.append(dict(zip(header, row_as_list)))

            # Handle each lot dict
            index_in_collection = 1
            for lot_dict in lots_data:

                # Catalogue and Collection
                collection_short_title = os.path.splitext(os.path.basename(transcription_file))[0]
                catalogue, created = Catalogue.objects.get_or_create(name=collection_short_title)
                collection, created = Collection.objects.get_or_create(
                    catalogue=catalogue,
                    short_title=collection_short_title,
                    full_title=lot_dict['full_collection_title'],
                    preface_and_paratexts=lot_dict['preface_and paratexts'],
                    year_of_publication=lot_dict['collection_date_of_publication'],
                )

                # Category
                category, created = Category.objects\
                    .get_or_create(collection=collection, bookseller_category=lot_dict['bookseller_category_books'])

                # Lot
                lot = Lot.objects.create(
                    collection=collection,
                    number_in_collection=lot_dict['number_in_collection'],
                    lot_as_listed_in_collection=lot_dict['item_as_listed_in_collection'],
                    index_in_collection=index_in_collection,
                    category=category
                )
                index_in_collection += 1

                # Place
                place = None
                if lot_dict['place_of_publication']:
                    try:
                        place = Place.objects.get(name=lot_dict['place_of_publication'])
                    except:
                        pass

                # Edition
                edition = Edition.objects.create(place=place)

                # Book format
                book_format, created = BookFormat.objects.get_or_create(name=lot_dict['book_format'])

                # Item
                item = Item.objects.create(
                    short_title=lot_dict['item_as_listed_in_collection'][:128],
                    lot=lot,
                    catalogue=catalogue,
                    book_format=book_format,
                    index_in_lot=1,
                    edition=edition
                )
